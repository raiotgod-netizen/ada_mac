"""
universal_tools.py — Universal PC Control Tool for ADA3.0

Single tool: universal_control(action, target, value)
Actions: sysinfo | cpu_detailed | gpu_info | memory_detailed | disk_smart | thermal_zones |
         net_connections | net_arp | net_routes | net_dns_cache | wifi_networks | wifi_connected |
         windows_full | window_tree | window_info | window_close | window_minimize |
         window_maximize | window_focus | window_set_pos | window_set_size | window_always_on_top |
         processes_detailed | process_tree | process_kill | process_kill_tree |
         dir_tree | file_search | file_info | file_duplicate_finder | file_watch |
         screenshot | screenshot_area | screen_regions | ocr_screen |
         clipboard_read | clipboard_write | key_send | mouse_move | mouse_click | mouse_drag |
         app_open | app_close | service_control | startup_list |
         env_get | env_set | registry_read | registry_write |
         reboot | shutdown | hibernate | lock |
         system_logs | crash_dump_check | update_status |
         excel_read | excel_write | browser_tabs
"""

from __future__ import annotations
import os
import sys
import subprocess
import base64
import tempfile
import hashlib
import time
import re
import shutil
from pathlib import Path
from typing import Optional, Dict, Any, List


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _run_ps(script: str, timeout: int = 10) -> subprocess.CompletedProcess:
    """Run PowerShell script, return result."""
    return subprocess.run(
        ["powershell", "-Command", script],
        capture_output=True, text=True, timeout=timeout
    )

def _run_cmd(cmd: List[str], timeout: int = 10) -> subprocess.CompletedProcess:
    """Run shell command."""
    return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)

# ─────────────────────────────────────────────────────────────
# SYSTEM
# ─────────────────────────────────────────────────────────────

def _sysinfo() -> str:
    import psutil, platform, datetime
    try:
        cpu = psutil.cpu_percent(interval=None)
        cpu_count = psutil.cpu_count()
        cpu_freq = psutil.cpu_freq()
        mem = psutil.virtual_memory()
        disk = psutil.disk_usage('C:\\')
        battery = psutil.sensors_battery()
        boot = datetime.datetime.fromtimestamp(psutil.boot_time())
        uptime = str(datetime.datetime.now() - boot).split('.')[0]
        freq_str = f"{round(cpu_freq.current/1000,1)}GHz" if cpu_freq else "N/A"
        battery_str = f"{battery.percent}% {'charging' if battery.power_plugged else 'on battery'}" if battery else "no battery"
        return (f"CPU: {cpu}% ({cpu_count} cores @ {freq_str})\n"
                f"RAM: {mem.percent}% ({mem.used//(1024**3)}/{mem.total//(1024**3)}GB)\n"
                f"Disk C: {round(disk.percent,1)}% ({disk.free//(1024**3)}GB free)\n"
                f"Battery: {battery_str}\nUptime: {uptime}")
    except Exception as e:
        return f"sysinfo error: {e}"

def _cpu_detailed() -> str:
    import psutil, platform
    try:
        cpu = psutil.cpu_percent(interval=None)
        per_cpu = psutil.cpu_percent(interval=None, percpu=True)
        freq = psutil.cpu_freq()
        ctx = psutil.cpu_stats()
        load = os.getloadavg() if hasattr(os, 'getloadavg') else [0,0,0]
        return (f"Overall: {cpu}%\nPer core: {per_cpu}\n"
                f"Freq: {round(freq.current/1000,1) if freq else 'N/A'}GHz\n"
                f"Ctx switches: {ctx.ctx_switches:,} | Interrupts: {ctx.interrupts:,}\n"
                f"Load (1m): {load[0]}")
    except Exception as e:
        return f"cpu_detailed error: {e}"

def _gpu_info() -> str:
    try:
        import pynvml
        pynvml.nvmlInit()
        handle = pynvml.nvmlDeviceGetHandleByIndex(0)
        name = pynvml.nvmlDeviceGetName(handle)
        mem_info = pynvml.nvmlDeviceGetMemoryInfo(handle)
        util = pynvml.nvmlDeviceGetUtilizationRates(handle)
        temp = pynvml.nvmlDeviceGetTemperatureHandle(handle, pynvml.NVML_TEMPERATURE_GPU)
        pynvml.nvmlShutdown()
        return (f"GPU: {name}\nVRAM: {mem_info.used//(1024**2)}MB / {mem_info.total//(1024**2)}MB\n"
                f"Util: {util.gpu}% | Temp: {temp}°C")
    except Exception as e:
        return f"gpu_info (NVIDIA only): {e}"

def _memory_detailed() -> str:
    import psutil
    try:
        vm = psutil.virtual_memory()
        swap = psutil.swap_memory()
        return (f"Virtual: {vm.percent}% | {vm.used//(1024**3)}GB/{vm.total//(1024**3)}GB\n"
                f"Available: {vm.available//(1024**3)}GB\n"
                f"Swap: {swap.percent}% | {swap.used//(1024**3)}GB/{swap.total//(1024**3)}GB")
    except Exception as e:
        return f"memory_detailed error: {e}"

def _disk_smart() -> str:
    try:
        result = _run_ps("Get-PhysicalDisk | Select-Object FriendlyName, MediaType, OperationalStatus, HealthStatus | ConvertTo-Json -Compress")
        return result.stdout.strip()[:500] if result.returncode == 0 else f"disk_smart error: {result.stderr[:200]}"
    except Exception as e:
        return f"disk_smart error: {e}"

def _thermal_zones() -> str:
    try:
        import psutil
        zones = psutil.sensors_temperatures()
        if not zones:
            return "No thermal data available"
        lines = []
        for name, entries in zones.items():
            for entry in entries:
                lines.append(f"{name}: {entry.label or 'temp'} = {entry.current}°C (high: {entry.high})")
        return "\n".join(lines[:20])
    except Exception as e:
        return f"thermal_zones error: {e}"

# ─────────────────────────────────────────────────────────────
# NETWORK
# ─────────────────────────────────────────────────────────────

def _net_connections() -> str:
    try:
        import psutil
        lines = []
        for conn in psutil.net_connections(kind='inet')[:50]:
            try:
                lines.append(f"{conn.laddr.ip}:{conn.laddr.port} -> {conn.raddr.ip if conn.raddr else '---'}:{conn.raddr.port if conn.raddr else '---'} [{conn.status}] pid:{conn.pid}")
            except Exception:
                pass
        return "\n".join(lines[:40]) or "No connections"
    except Exception as e:
        return f"net_connections error: {e}"

def _net_arp() -> str:
    result = _run_ps("Get-NetNeighbor | Where-Object { $_.State -ne 'Unreachable' } | Select-Object IPAddress, LinkLayerAddress | ConvertTo-Json -Compress")
    return result.stdout.strip()[:800] if result.returncode == 0 else f"arp error: {result.stderr[:200]}"

def _net_routes() -> str:
    result = _run_ps("route print | Select-Object -First 30")
    return result.stdout.strip()[:800] if result.returncode == 0 else f"routes error: {result.stderr[:200]}"

def _net_dns_cache() -> str:
    result = _run_ps("Get-DnsClientCache | Select-Object -First 20 Entry, Data | ConvertTo-Json -Compress")
    return result.stdout.strip()[:600] if result.returncode == 0 else f"dns cache error: {result.stderr[:200]}"

def _wifi_networks() -> str:
    result = _run_ps("netsh wlan show networks mode=bssid | Select-Object -First 40")
    return result.stdout.strip()[:600] if result.returncode == 0 else f"wifi error: {result.stderr[:200]}"

def _wifi_connected() -> str:
    result = _run_ps("netsh wlan show interfaces | Select-Object -First 15")
    return result.stdout.strip()[:600] if result.returncode == 0 else f"wifi error: {result.stderr[:200]}"

def _firewall_rules() -> str:
    result = _run_ps("Get-NetFirewallRule | Where-Object { $_.Enabled -eq $true } | Select-Object -First 20 Name, Direction, Action | ConvertTo-Json -Compress")
    return result.stdout.strip()[:800] if result.returncode == 0 else f"firewall error: {result.stderr[:200]}"

# ─────────────────────────────────────────────────────────────
# WINDOWS
# ─────────────────────────────────────────────────────────────

def _windows_full() -> str:
    try:
        import win32gui
        result = []
        def callback(hwnd, data):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                if title:
                    cl_name = win32gui.GetClassName(hwnd)
                    rect = win32gui.GetWindowRect(hwnd)
                    result.append(f"[{hwnd}] {title[:60]} ({cl_name}) rect:{rect}")
        win32gui.EnumWindows(callback, None)
        return "\n".join(result[:30])
    except Exception as e:
        return f"windows_full error: {e}"

def _window_tree() -> str:
    try:
        import win32gui
        result = []
        def callback(hwnd, data):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)[:50]
                result.append(f"{'  '*data['depth']}{title}")
                data['depth'] += 1
                win32gui.EnumChildWindows(hwnd, callback, data)
                data['depth'] -= 1
        win32gui.EnumWindows(callback, {'depth': 0})
        return "\n".join(result[:40])
    except Exception as e:
        return f"window_tree error: {e}"

def _window_info(target: str) -> str:
    try:
        import win32gui
        def callback(hwnd, wins):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                if target.lower() in title.lower():
                    wins.append({'hwnd': hwnd, 'title': title, 'class': win32gui.GetClassName(hwnd), 'rect': win32gui.GetWindowRect(hwnd)})
        wins = []
        win32gui.EnumWindows(callback, wins)
        if not wins:
            return f"No window matching: {target}"
        w = wins[0]
        return f"Title: {w['title']}\nClass: {w['class']}\nHWND: {w['hwnd']}\nRect: {w['rect']}"
    except Exception as e:
        return f"window_info error: {e}"

def _window_action(target: str, action: str) -> str:
    try:
        import win32gui, win32con, win32api
        def callback(hwnd, wins):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                if target.lower() in title.lower():
                    wins.append(hwnd)
        wins = []
        win32gui.EnumWindows(callback, wins)
        if not wins:
            return f"No window matching: {target}"
        hwnd = wins[0]
        if action == "close":
            win32api.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
            return f"Closed: {target}"
        elif action == "minimize":
            win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)
            return f"Minimized: {target}"
        elif action == "maximize":
            win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
            return f"Maximized: {target}"
        elif action == "focus":
            win32gui.SetForegroundWindow(hwnd)
            return f"Focused: {target}"
        return f"Unknown action: {action}"
    except Exception as e:
        return f"window_action error: {e}"

def _window_set_pos(target: str, x: int, y: int, w: int = None, h: int = None) -> str:
    try:
        import win32gui, win32con
        def callback(hwnd, wins):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                if target.lower() in title.lower():
                    wins.append(hwnd)
        wins = []
        win32gui.EnumWindows(callback, wins)
        if not wins:
            return f"No window matching: {target}"
        hwnd = wins[0]
        if w and h:
            win32gui.MoveWindow(hwnd, x, y, w, h, True)
        else:
            win32gui.SetWindowPos(hwnd, 0, x, y, 0, 0, win32con.SWP_NOSIZE)
        return f"Moved {target} to ({x},{y})"
    except Exception as e:
        return f"window_set_pos error: {e}"

def _window_always_on_top(target: str, enable: bool = True) -> str:
    try:
        import win32gui, win32con
        def callback(hwnd, wins):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                if target.lower() in title.lower():
                    wins.append(hwnd)
        wins = []
        win32gui.EnumWindows(callback, wins)
        if not wins:
            return f"No window matching: {target}"
        hwnd = wins[0]
        flag = win32con.HWND_TOPMOST if enable else win32con.HWND_NOTOPMOST
        win32gui.SetWindowPos(hwnd, flag, 0, 0, 0, 0, win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)
        return f"{'Enabled' if enable else 'Disabled'} always-on-top for {target}"
    except Exception as e:
        return f"window_always_on_top error: {e}"

# ─────────────────────────────────────────────────────────────
# PROCESSES
# ─────────────────────────────────────────────────────────────

def _processes_detailed() -> str:
    try:
        import psutil
        lines = []
        for p in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_percent', 'num_handles', 'num_threads']):
            try:
                info = p.info
                lines.append(f"PID:{info['pid']} {info['name'][:30]:30} CPU:{info['cpu_percent']}% MEM:{round(info['memory_percent'],1)}% HNDLS:{info['num_handles']} TH:{info['num_threads']}")
            except Exception:
                pass
        return "\n".join(lines[:40])
    except Exception as e:
        return f"processes error: {e}"

def _process_tree() -> str:
    try:
        import psutil
        lines = []
        for p in psutil.process_iter(['pid', 'name', 'ppid']):
            try:
                info = p.info
                lines.append(f"PID:{info['pid']} PPID:{info['ppid']} {info['name']}")
            except Exception:
                pass
        return "\n".join(lines[:60])
    except Exception as e:
        return f"process_tree error: {e}"

def _process_kill(target: str) -> str:
    try:
        import psutil, subprocess
        killed = []
        for p in psutil.process_iter(['pid', 'name']):
            try:
                if target.lower() in p.info['name'].lower():
                    pid = p.info['pid']
                    subprocess.run(['taskkill', '/PID', str(pid), '/F'], capture_output=True)
                    killed.append(str(pid))
            except Exception:
                pass
        return f"Killed PIDs: {', '.join(killed) if killed else 'none found'}"
    except Exception as e:
        return f"process_kill error: {e}"

def _process_kill_tree(target: str) -> str:
    try:
        import psutil, subprocess
        killed = []
        for p in psutil.process_iter(['pid', 'name', 'ppid']):
            try:
                if target.lower() in p.info['name'].lower():
                    pid = p.info['pid']
                    subprocess.run(['taskkill', '/PID', str(pid), '/T', '/F'], capture_output=True)
                    killed.append(str(pid))
            except Exception:
                pass
        return f"Killed tree PIDs: {', '.join(killed) if killed else 'none found'}"
    except Exception as e:
        return f"process_kill_tree error: {e}"

# ─────────────────────────────────────────────────────────────
# FILES
# ─────────────────────────────────────────────────────────────

def _dir_tree(path: str = None, depth: int = 3) -> str:
    path = path or os.path.expanduser("~")
    try:
        lines = []
        for root, dirs, files in os.walk(path):
            depth_dir = root.replace(path, '').count(os.sep)
            if depth_dir >= depth:
                dirs.clear()
                continue
            lines.append(f"{'  '*depth_dir}{os.path.basename(root)}/")
            for f in files[:5]:
                lines.append(f"{'  '*(depth_dir+1)}{f}")
        return "\n".join(lines[:80])
    except Exception as e:
        return f"dir_tree error: {e}"

def _file_search(pattern: str, path: str = None) -> str:
    path = path or os.path.expanduser("~")
    try:
        matches = []
        for root, dirs, files in os.walk(path):
            for f in files:
                if pattern.lower() in f.lower():
                    matches.append(os.path.join(root, f))
            if len(matches) >= 30:
                break
        return "\n".join(matches[:30])
    except Exception as e:
        return f"file_search error: {e}"

def _file_info(path: str) -> str:
    try:
        p = Path(path)
        if not p.exists():
            return f"File not found: {path}"
        stat = p.stat()
        attrs = os.attrs(p) if hasattr(os, 'attrs') else {}
        return (f"Path: {p.absolute()}\n"
                f"Size: {stat.st_size:,} bytes\n"
                f"Created: {time.ctime(stat.st_ctime)}\n"
                f"Modified: {time.ctime(stat.st_mtime)}\n"
                f"Accessed: {time.ctime(stat.st_atime)}\n"
                f"Attributes: {attrs}")
    except Exception as e:
        return f"file_info error: {e}"

def _file_duplicate_finder(path: str = None) -> str:
    path = path or os.path.expanduser("~")
    try:
        hashes = {}
        dups = []
        for root, dirs, files in os.walk(path):
            for f in files:
                fp = os.path.join(root, f)
                try:
                    size = os.path.getsize(fp)
                    if size < 1024:
                        continue
                    with open(fp, 'rb') as fh:
                        h = hashlib.md5(fh.read(4096)).hexdigest()
                    key = (size, h)
                    if key in hashes:
                        dups.append(f"{fp} == {hashes[key]}")
                    else:
                        hashes[key] = fp
                except Exception:
                    pass
                if len(dups) >= 20:
                    break
            if len(dups) >= 20:
                break
        return "\n".join(dups) if dups else "No duplicates found"
    except Exception as e:
        return f"file_duplicate_finder error: {e}"

def _file_watch(path: str, seconds: int = 5) -> str:
    try:
        import watchdog
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
        class Handler(FileSystemEventHandler):
            def __init__(self):
                self.events = []
            def on_any_event(self, event):
                if not event.is_directory:
                    self.events.append(f"{event.event_type}: {event.src_path}")
        h = Handler()
        Observer().schedule(h, path, recursive=True).start()
        time.sleep(seconds)
        Observer().stop()
        return "\n".join(h.events[:20]) if h.events else "No changes detected"
    except Exception as e:
        return f"file_watch error: {e}"

def _file_delete(path: str) -> str:
    try:
        p = Path(path)
        if p.is_file():
            p.unlink()
            return f"Deleted: {path}"
        elif p.is_dir():
            shutil.rmtree(p)
            return f"Deleted dir: {path}"
        return f"Not found: {path}"
    except Exception as e:
        return f"file_delete error: {e}"

# ─────────────────────────────────────────────────────────────
# SCREENSHOT & VISION
# ─────────────────────────────────────────────────────────────

def _screenshot() -> str:
    try:
        import mss
        with mss.mss() as sct:
            monitor = sct.monitors[1]
            screenshot = sct.grab(monitor)
            img_bytes = mss.tools.to_png(screenshot.rgb, screenshot.size)
            b64 = base64.b64encode(img_bytes).decode('utf-8')
            return f"[SCREENSHOT_BASE64:{len(b64)} chars]"
    except Exception as e:
        return f"screenshot error: {e}"

def _screenshot_area(x: int, y: int, w: int, h: int) -> str:
    try:
        import mss
        with mss.mss() as sct:
            monitor = sct.monitors[1]
            bbox = (x, y, x+w, y+h)
            screenshot = sct.grab(bbox)
            img_bytes = mss.tools.to_png(screenshot.rgb, screenshot.size)
            b64 = base64.b64encode(img_bytes).decode('utf-8')
            return f"[SCREENSHOT_AREA_BASE64:{len(b64)} chars]"
    except Exception as e:
        return f"screenshot_area error: {e}"

def _screen_regions() -> str:
    try:
        import mss
        with mss.mss() as sct:
            monitor = sct.monitors[1]
            screenshot = sct.grab(monitor)
            # Simple: just return monitor info
            return f"Monitor: {monitor} size: {screenshot.size}"
    except Exception as e:
        return f"screen_regions error: {e}"

def _ocr_screen() -> str:
    try:
        import mss
        with mss.mss() as sct:
            monitor = sct.monitors[1]
            screenshot = sct.grab(monitor)
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
                f.write(mss.tools.to_png(screenshot.rgb, screenshot.size))
                temp_path = f.name
        from ocr_engine import OCREngine
        ocr = OCREngine()
        result = ocr.extract(temp_path)
        os.unlink(temp_path)
        if result.get("ocr_ready"):
            lines = [l['text'] for l in result.get("lines", []) if l.get('text', '').strip()]
            return " | ".join(lines[:30])
        return "No text found"
    except Exception as e:
        return f"ocr_screen error: {e}"

# ─────────────────────────────────────────────────────────────
# CLIPBOARD & INPUT
# ─────────────────────────────────────────────────────────────

def _clipboard_read() -> str:
    try:
        import win32clipboard
        win32clipboard.OpenClipboard()
        if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_TEXT):
            data = win32clipboard.GetClipboardData(win32clipboard.CF_TEXT).decode('utf-8', errors='replace')
        elif win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_UNICODETEXT):
            data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
        else:
            data = "(clipboard empty or unsupported format)"
        win32clipboard.CloseClipboard()
        return data[:500]
    except Exception as e:
        return f"clipboard_read error: {e}"

def _clipboard_write(text: str) -> str:
    try:
        import win32clipboard
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(text)
        win32clipboard.CloseClipboard()
        return f"Written to clipboard ({len(text)} chars)"
    except Exception as e:
        return f"clipboard_write error: {e}"

def _key_send(keys: str) -> str:
    try:
        import pyautogui
        pyautogui.write(keys, interval=0.02)
        return f"Sent keys: {keys[:50]}"
    except Exception as e:
        return f"key_send error: {e}"

def _mouse_move(x: int, y: int) -> str:
    try:
        import pyautogui
        pyautogui.moveTo(x, y, duration=0.2)
        return f"Moved mouse to ({x},{y})"
    except Exception as e:
        return f"mouse_move error: {e}"

def _mouse_click(x: int = None, y: int = None, button: str = "left") -> str:
    try:
        import pyautogui
        if x is not None and y is not None:
            pyautogui.click(x, y, button=button)
        else:
            pyautogui.click(button=button)
        return f"Clicked {button} at ({x},{y})" if x else f"Clicked {button}"
    except Exception as e:
        return f"mouse_click error: {e}"

def _mouse_drag(x1: int, y1: int, x2: int, y2: int) -> str:
    try:
        import pyautogui
        pyautogui.moveTo(x1, y1)
        pyautogui.drag(x2-x1, y2-y1, duration=0.5)
        return f"Dragged from ({x1},{y1}) to ({x2},{y2})"
    except Exception as e:
        return f"mouse_drag error: {e}"

def _press_hotkey(keys: str) -> str:
    """Press a hotkey combination. Keys separated by '+', e.g. 'ctrl+c', 'alt+tab', 'ctrl+shift+esc'."""
    try:
        import pyautogui
        key_list = [k.strip().lower() for k in keys.split("+")]
        # pyautogui needs hotkey(*keys) where keys are lowercase
        pyautogui.hotkey(*key_list)
        return f"Pressed hotkey: {'+'.join(key_list)}"
    except Exception as e:
        return f"press_hotkey error: {e}"

def _scroll(clicks: int, axis: str = "vertical") -> str:
    """Scroll the mouse. axis='vertical' or 'horizontal'. Use negative for opposite direction."""
    try:
        import pyautogui
        axis = axis.lower().strip()
        if axis == "horizontal" or axis == "h":
            pyautogui.hscroll(clicks)
        else:
            pyautogui.scroll(clicks)
        return f"Scrolled {axis} {clicks} clicks"
    except Exception as e:
        return f"scroll error: {e}"

def _type_text(text: str, interval: float = None) -> str:
    """Type text with optional interval between keystrokes. Value param = interval."""
    try:
        import pyautogui
        intv = float(value) if value else 0.02
        pyautogui.write(text, interval=intv)
        return f"Typed {len(text)} characters"
    except Exception as e:
        return f"type_text error: {e}"

# ─────────────────────────────────────────────────────────────
# APP CONTROL
# ─────────────────────────────────────────────────────────────

def _app_open(target: str) -> str:
    try:
        # URL
        if target.startswith('http://') or target.startswith('https://'):
            os.startfile(target)
            return f"Opened URL: {target}"
        # File path
        if os.path.exists(target):
            os.startfile(target)
            return f"Opened: {target}"
        # Common apps
        app_paths = {
            "spotify": os.path.expandvars(r"%LOCALAPPDATA%\Spotify\Spotify.exe"),
            "discord": os.path.expandvars(r"%LOCALAPPDATA%\Discord\Update.exe"),
            "vscode": os.path.expandvars(r"%LOCALAPPDATA%\Programs\Microsoft VS Code\Code.exe"),
            "notepad": os.path.expandvars(r"%SystemRoot%\System32\notepad.exe"),
            "calculator": os.path.expandvars(r"%SystemRoot%\System32\calc.exe"),
            "explorer": os.path.expandvars(r"%SystemRoot%\explorer.exe"),
            "chrome": os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
            "edge": os.path.expandvars(r"%ProgramFiles (x86)%\Microsoft\Edge\Application\msedge.exe"),
            "teams": os.path.expandvars(r"%AppData%\Microsoft\Teams\Update.exe"),
        }
        target_lower = target.lower().strip()
        if target_lower in app_paths:
            path = app_paths[target_lower]
            if os.path.exists(path):
                os.startfile(path)
                return f"Opened: {target}"
        # Try shell
        subprocess.run(["cmd", "/c", "start", "", target], capture_output=True, timeout=5)
        return f"Attempted: {target}"
    except Exception as e:
        return f"app_open error: {e}"

def _app_close(target: str) -> str:
    try:
        import psutil, subprocess
        killed = []
        for p in psutil.process_iter(['pid', 'name']):
            try:
                if target.lower() in p.info['name'].lower():
                    subprocess.run(['taskkill', '/PID', str(p.info['pid']), '/F'], capture_output=True)
                    killed.append(p.info['name'])
            except Exception:
                pass
        return f"Closed: {', '.join(killed) if killed else 'none found'}"
    except Exception as e:
        return f"app_close error: {e}"

# ─────────────────────────────────────────────────────────────
# SERVICES
# ─────────────────────────────────────────────────────────────

def _service_control(service_name: str, action: str) -> str:
    valid = ["start", "stop", "restart", "pause", "resume"]
    if action not in valid:
        return f"Invalid action. Use: {', '.join(valid)}"
    try:
        result = _run_ps(f"sc {action} {service_name}")
        return f"Service {service_name} {action}: {'OK' if result.returncode == 0 else result.stderr[:200]}"
    except Exception as e:
        return f"service_control error: {e}"

def _startup_list() -> str:
    result = _run_ps("Get-CimInstance -ClassName Win32_StartupCommand | Select-Object Name, Command, Location | ConvertTo-Json -Compress")
    return result.stdout.strip()[:800] if result.returncode == 0 else f"startup_list error: {result.stderr[:200]}"

# ─────────────────────────────────────────────────────────────
# ENVIRONMENT & REGISTRY
# ─────────────────────────────────────────────────────────────

def _env_get(var: str = None) -> str:
    if var:
        return os.environ.get(var, "(not set)")
    return "\n".join([f"{k}={v}" for k, v in os.environ.items()][:30])

def _env_set(var: str, value: str) -> str:
    try:
        os.environ[var] = value
        return f"Set {var}={value}"
    except Exception as e:
        return f"env_set error: {e}"

def _registry_read(path: str) -> str:
    result = _run_ps(f"Get-ItemProperty -Path '{path}' | ConvertTo-Json -Compress", timeout=5)
    return result.stdout.strip()[:600] if result.returncode == 0 else f"registry_read error: {result.stderr[:200]}"

def _registry_write(path: str, name: str, value: str) -> str:
    result = _run_ps(f"Set-ItemProperty -Path '{path}' -Name '{name}' -Value '{value}'", timeout=5)
    return f"Registry set: {path}\\{name}={value}" if result.returncode == 0 else f"registry_write error: {result.stderr[:200]}"

# ─────────────────────────────────────────────────────────────
# PC CONTROL
# ─────────────────────────────────────────────────────────────

def _reboot() -> str:
    subprocess.run(["shutdown", "/r", "/t", "10", "/c", "Reboot initiated by ADA"], capture_output=True)
    return "Reboot scheduled in 10 seconds"

def _shutdown() -> str:
    subprocess.run(["shutdown", "/s", "/t", "10", "/c", "Shutdown initiated by ADA"], capture_output=True)
    return "Shutdown scheduled in 10 seconds"

def _hibernate() -> str:
    subprocess.run(["shutdown", "/h"], capture_output=True)
    return "Hibernation initiated"

def _lock() -> str:
    subprocess.run(["rundll32.exe", "user32.dll,LockWorkStation"], capture_output=True)
    return "Workstation locked"

# ─────────────────────────────────────────────────────────────
# SYSTEM LOGS & STATUS
# ─────────────────────────────────────────────────────────────

def _system_logs(lines: int = 20) -> str:
    result = _run_ps(f"Get-EventLog -LogName System -Newest {lines} | Select-Object TimeGenerated, EntryType, Source, Message | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:800] if result.returncode == 0 else f"system_logs error: {result.stderr[:200]}"

def _crash_dump_check() -> str:
    crash_path = os.path.expandvars(r"%SystemRoot%\Minidump")
    if os.path.exists(crash_path):
        files = os.listdir(crash_path)[:5]
        total = sum(os.path.getsize(os.path.join(crash_path, f)) for f in files)
        return f"Minidumps: {len(files)} files, {total//(1024*1024)}MB in {crash_path}"
    return "No minidumps found"

def _update_status() -> str:
    result = _run_ps("Get-Service wuauserv | Select-Object Status, StartType | ConvertTo-Json -Compress", timeout=5)
    return result.stdout.strip()[:400] if result.returncode == 0 else f"update_status error: {result.stderr[:200]}"

# ─────────────────────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────────────────────

def _excel_read(path: str, sheet: str = None, cell_range: str = None) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        if cell_range:
            data = [[str(cell.value or '') for cell in row] for row in ws[cell_range]]
        else:
            data = [[str(cell.value or '') for cell in row] for row in ws.iter_rows(max_row=20, max_col=10)]
        return "\n".join([" | ".join(row) for row in data])
    except Exception as e:
        return f"excel_read error: {e}"

def _excel_write(path: str, cell: str, value: str, sheet: str = None) -> str:
    try:
        import openpyxl
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
        ws = wb[sheet] if sheet else wb.active
        ws[cell] = value
        wb.save(path)
        return f"Written {value} to {cell} in {path}"
    except Exception as e:
        return f"excel_write error: {e}"

# ─────────────────────────────────────────────────────────────
# BROWSER TABS
# ─────────────────────────────────────────────────────────────

def _browser_tabs() -> str:
    try:
        import win32gui
        result = []
        def callback(hwnd, wins):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                if title and any(b in title for b in ["Chrome", "Edge", "Firefox", "Opera"]):
                    result.append(title[:80])
        win32gui.EnumWindows(callback, None)
        return "\n".join(result[:20])
    except Exception as e:
        return f"browser_tabs error: {e}"

# ─────────────────────────────────────────────────────────────
# WMI & POWERSHELL ADVANCED
# ─────────────────────────────────────────────────────────────

def _wmi_query(query: str) -> str:
    result = _run_ps(f"Get-WmiObject -Query \"{query}\" | ConvertTo-Json -Compress", timeout=15)
    return result.stdout.strip()[:1000] if result.returncode == 0 else f"wmi_query error: {result.stderr[:200]}"

def _powershell_eval(script: str) -> str:
    result = _run_ps(script, timeout=20)
    out = result.stdout.strip()[:800]
    err = result.stderr.strip()[:200]
    return f"Output: {out}\nError: {err}" if err else f"Output: {out}"

def _event_log(log_name: str = "System", lines: int = 20) -> str:
    result = _run_ps(f"Get-EventLog -LogName {log_name} -Newest {lines} | Select-Object TimeGenerated, EntryType, Source, Message | ConvertTo-Json -Compress", timeout=15)
    return result.stdout.strip()[:1000] if result.returncode == 0 else f"event_log error: {result.stderr[:200]}"

def _scheduled_tasks() -> str:
    result = _run_ps("Get-ScheduledTask | Select-Object TaskName, State, TaskPath | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:800] if result.returncode == 0 else f"scheduled_tasks error: {result.stderr[:200]}"

# ─────────────────────────────────────────────────────────────
# HARDWARE & SENSORS
# ─────────────────────────────────────────────────────────────

def _fan_speed() -> str:
    result = _run_ps("Get-WmiObject MSAcpi_ThermalZoneTemperature -Namespace root/wmi | Select-Object InstanceName, CurrentTemperature | ConvertTo-Json -Compress", timeout=5)
    return result.stdout.strip()[:600] if result.returncode == 0 else f"fan_speed (WMI): {result.stderr[:200]}"

def _disk_partitions() -> str:
    result = _run_ps("Get-WmiObject -Class Win32_DiskPartition | Select-Object Name, Size, Bootable, BootPartition, PrimaryPartition | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:800] if result.returncode == 0 else f"disk_partitions error: {result.stderr[:200]}"

def _hardware_usb() -> str:
    result = _run_ps("Get-PnpDevice -Class USB -Status OK | Select-Object FriendlyName, Status | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:800] if result.returncode == 0 else f"hardware_usb error: {result.stderr[:200]}"

def _bluetooth_devices() -> str:
    result = _run_ps("Get-PnpDevice -Class Bluetooth -Status OK | Select-Object FriendlyName | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:600] if result.returncode == 0 else f"bluetooth_devices error: {result.stderr[:200]}"

# ─────────────────────────────────────────────────────────────
# NETWORK ADVANCED
# ─────────────────────────────────────────────────────────────

def _ping_status(host: str) -> str:
    result = _run_ps(f"Test-Connection -ComputerName {host} -Count 2 | Select-Object Address, Latency, Status | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:600] if result.returncode == 0 else f"ping error: {result.stderr[:200]}"

def _port_scan(port_range: str = "1-1024") -> str:
    try:
        start, end = port_range.split("-")
        result = _run_ps(f"PortTest -ComputerName localhost -StartPort {start} -EndPort {end} | ConvertTo-Json -Compress", timeout=15)
        return result.stdout.strip()[:800] if result.returncode == 0 else f"Use: netstat -an | Select-String LISTENING"
    except Exception as e:
        return f"port_scan: use netstat approach instead"

def _network_map() -> str:
    result = _run_ps("Get-NetNeighbor | Where-Object { $_.State -ne 'Unreachable' } | Select-Object IPAddress, LinkLayerAddress | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:800] if result.returncode == 0 else f"network_map error: {result.stderr[:200]}"

# ─────────────────────────────────────────────────────────────
# APPLICATIONS ADVANCED
# ─────────────────────────────────────────────────────────────

def _pdf_read(path: str) -> str:
    try:
        import PyPDF2
        with open(path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            text = ""
            for page in reader.pages[:5]:
                text += page.extract_text() or ""
            return text[:500] if text.strip() else "No text found in PDF"
    except ImportError:
        return "PyPDF2 not installed. Install with: pip install PyPDF2"
    except Exception as e:
        return f"pdf_read error: {e}"

def _compress_file(source: str, destination: str = None) -> str:
    try:
        import zipfile
        if not destination:
            destination = source + ".zip"
        with zipfile.ZipFile(destination, 'w', zipfile.ZIP_DEFLATED) as zf:
            if os.path.isdir(source):
                for root, dirs, files in os.walk(source):
                    for file in files:
                        zf.write(os.path.join(root, file))
            else:
                zf.write(source)
        return f"Compressed: {destination}"
    except Exception as e:
        return f"compress_file error: {e}"

def _uncompress_file(source: str, destination: str = None) -> str:
    try:
        import zipfile
        if destination is None:
            destination = source.replace('.zip', '_unzipped')
        with zipfile.ZipFile(source, 'r') as zf:
            zf.extractall(destination)
        return f"Extracted to: {destination}"
    except Exception as e:
        return f"uncompress_file error: {e}"

# ─────────────────────────────────────────────────────────────
# SECURITY
# ─────────────────────────────────────────────────────────────

def _firewall_status() -> str:
    result = _run_ps("Get-NetFirewallProfile | Select-Object Name, Enabled | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:600] if result.returncode == 0 else f"firewall_status error: {result.stderr[:200]}"

def _antivirus_status() -> str:
    result = _run_ps("Get-MpComputerStatus | Select-Object AntivirusEnabled, RealTimeProtectionEnabled, AntivirusSignatureLastUpdated | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:600] if result.returncode == 0 else f"antivirus_status error: {result.stderr[:200]}"

def _uac_check() -> str:
    result = _run_ps("(Get-ItemProperty -Path HKLM:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System).EnableLUA", timeout=5)
    enabled = "1" in result.stdout
    return f"UAC: {'ENABLED' if enabled else 'DISABLED'}"

# ─────────────────────────────────────────────────────────────
# ADVANCED DEBUG
# ─────────────────────────────────────────────────────────────

def _process_memory(pid: int = None) -> str:
    try:
        import psutil
        if pid:
            p = psutil.Process(pid)
            return f"PID {pid}: {p.name()} - MEM: {p.memory_info().rss//(1024**2)}MB"
        lines = []
        for p in psutil.process_iter(['pid', 'name', 'memory_info']):
            try:
                info = p.info
                lines.append(f"PID:{info['pid']} {info['name'][:30]:30} MEM:{info['memory_info'].rss//(1024**2):6}MB")
            except Exception:
                pass
        return "\n".join(lines[:30])
    except Exception as e:
        return f"process_memory error: {e}"

def _debug_output(lines: int = 20) -> str:
    result = _run_ps(f"Get-WinEvent -LogName 'System' -MaxEvents {lines} | Where-Object {{$_.LevelDisplayName -eq 'Error'}} | Select-Object TimeCreated, Message | ConvertTo-Json -Compress", timeout=10)
    return result.stdout.strip()[:1000] if result.returncode == 0 else f"debug_output error: {result.stderr[:200]}"

# ─────────────────────────────────────────────────────────────
# DATABASES
# ─────────────────────────────────────────────────────────────

def _sqlite_query(db_path: str, query: str) -> str:
    try:
        import sqlite3
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute(query)
        rows = cur.fetchall()
        conn.close()
        if not rows:
            return "No results"
        return "\n".join([" | ".join(str(c) for c in r) for r in rows[:50]])
    except Exception as e:
        return f"sqlite_query error: {e}"

# ─────────────────────────────────────────────────────────────
# MAIN DISPATCHER
# ─────────────────────────────────────────────────────────────

def universal_control(action: str, target: str = None, value: str = None) -> Dict[str, Any]:
    """
    Universal PC control tool.
    action: The action to perform (see action list above)
    target: Target entity (window name, file path, etc.)
    value: Additional parameter for the action
    """
    action = action.lower().strip() if action else ""

    # System
    if action == "sysinfo":          return {"result": _sysinfo()}
    if action == "cpu_detailed":    return {"result": _cpu_detailed()}
    if action == "gpu_info":        return {"result": _gpu_info()}
    if action == "memory_detailed": return {"result": _memory_detailed()}
    if action == "disk_smart":     return {"result": _disk_smart()}
    if action == "thermal_zones":   return {"result": _thermal_zones()}

    # Network
    if action == "net_connections": return {"result": _net_connections()}
    if action == "net_arp":         return {"result": _net_arp()}
    if action == "net_routes":      return {"result": _net_routes()}
    if action == "net_dns_cache":   return {"result": _net_dns_cache()}
    if action == "wifi_networks":   return {"result": _wifi_networks()}
    if action == "wifi_connected":  return {"result": _wifi_connected()}
    if action == "firewall_rules":  return {"result": _firewall_rules()}

    # Windows
    if action == "windows_full":          return {"result": _windows_full()}
    if action == "window_tree":          return {"result": _window_tree()}
    if action == "window_info":          return {"result": _window_info(target or "")}
    if action in ("window_close", "window_minimize", "window_maximize", "window_focus"):
        return {"result": _window_action(target or "", action.split("_")[1])}
    if action == "window_set_pos":
        parts = (target or "").split(",")
        x, y, w, h = int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else None, int(parts[3]) if len(parts) > 3 else None
        return {"result": _window_set_pos(value or "", x, y, w, h)}
    if action == "window_always_on_top":
        return {"result": _window_always_on_top(target or "", value == "true" or value == "1")}

    # Processes
    if action == "processes_detailed": return {"result": _processes_detailed()}
    if action == "process_tree":       return {"result": _process_tree()}
    if action == "process_kill":      return {"result": _process_kill(target or "")}
    if action == "process_kill_tree": return {"result": _process_kill_tree(target or "")}

    # Files
    if action == "dir_tree":          return {"result": _dir_tree(target)}
    if action == "file_search":       return {"result": _file_search(target or "~")}
    if action == "file_info":         return {"result": _file_info(target or "")}
    if action == "file_duplicate_finder": return {"result": _file_duplicate_finder(target)}
    if action == "file_watch":         return {"result": _file_watch(target or os.path.expanduser("~"), int(value) if value else 5)}
    if action == "file_delete":       return {"result": _file_delete(target or "")}

    # Screenshot & Vision
    if action == "screenshot":       return {"result": _screenshot()}
    if action == "screenshot_area":
        parts = (target or "0,0,100,100").split(",")
        x, y, w, h = int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3])
        return {"result": _screenshot_area(x, y, w, h)}
    if action == "screen_regions":   return {"result": _screen_regions()}
    if action == "ocr_screen":       return {"result": _ocr_screen()}

    # Clipboard & Input
    if action == "clipboard_read":   return {"result": _clipboard_read()}
    if action == "clipboard_write":  return {"result": _clipboard_write(target or "")}
    if action == "key_send":         return {"result": _key_send(target or "")}
    if action == "mouse_move":
        parts = (target or "0,0").split(",")
        return {"result": _mouse_move(int(parts[0]), int(parts[1]))}
    if action == "mouse_click":
        parts = (target or ",").split(",")
        x = int(parts[0]) if parts[0] else None
        y = int(parts[1]) if len(parts) > 1 and parts[1] else None
        return {"result": _mouse_click(x, y, value or "left")}
    if action == "mouse_drag":
        parts = (target or "0,0,0,0").split(",")
        return {"result": _mouse_drag(int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3]))}

    # App Control
    if action == "app_open":         return {"result": _app_open(target or "")}
    if action == "app_close":        return {"result": _app_close(target or "")}

    # Services
    if action == "service_control":
        parts = (target or ",").split(",")
        return {"result": _service_control(parts[0], parts[1] if len(parts) > 1 else "start")}
    if action == "startup_list":     return {"result": _startup_list()}

    # Environment & Registry
    if action == "env_get":          return {"result": _env_get(target)}
    if action == "env_set":          return {"result": _env_set(target or "", value or "")}
    if action == "registry_read":    return {"result": _registry_read(target or "")}
    if action == "registry_write":
        parts = (target or "").split(",")
        return {"result": _registry_write(parts[0], parts[1] if len(parts) > 1 else "", value or "")}

    # PC Control
    if action == "reboot":           return {"result": _reboot()}
    if action == "shutdown":         return {"result": _shutdown()}
    if action == "hibernate":         return {"result": _hibernate()}
    if action == "lock":             return {"result": _lock()}

    # System Logs
    if action == "system_logs":      return {"result": _system_logs(int(target) if target else 20)}
    if action == "crash_dump_check": return {"result": _crash_dump_check()}
    if action == "update_status":    return {"result": _update_status()}

    # Excel
    if action == "excel_read":       return {"result": _excel_read(target or "", value)}
    if action == "excel_write":
        parts = (target or "").split(",")
        return {"result": _excel_write(parts[0], parts[1] if len(parts) > 1 else "A1", value or "")}

    # Browser
    if action == "browser_tabs":     return {"result": _browser_tabs()}

    # WMI & PowerShell Advanced
    if action == "wmi_query":         return {"result": _wmi_query(target or "SELECT * FROM Win32_OperatingSystem")}
    if action == "powershell_eval":  return {"result": _powershell_eval(target or "Get-Process")}
    if action == "event_log":        return {"result": _event_log(target, int(value) if value else 20)}
    if action == "scheduled_tasks":  return {"result": _scheduled_tasks()}

    # Hardware & Sensors
    if action == "fan_speed":        return {"result": _fan_speed()}
    if action == "disk_partitions": return {"result": _disk_partitions()}
    if action == "hardware_usb":     return {"result": _hardware_usb()}
    if action == "bluetooth_devices":return {"result": _bluetooth_devices()}

    # Network Advanced
    if action == "ping_status":      return {"result": _ping_status(target or "google.com")}
    if action == "port_scan":       return {"result": _port_scan(target or "1-1024")}
    if action == "network_map":      return {"result": _network_map()}

    # Applications Advanced
    if action == "pdf_read":         return {"result": _pdf_read(target or "")}
    if action == "compress_file":     return {"result": _compress_file(target or "", value)}
    if action == "uncompress_file":  return {"result": _uncompress_file(target or "", value)}

    # Security
    if action == "firewall_status":   return {"result": _firewall_status()}
    if action == "antivirus_status": return {"result": _antivirus_status()}
    if action == "uac_check":         return {"result": _uac_check()}

    # Advanced Debug
    if action == "process_memory":
        pid = int(target) if target and target.isdigit() else None
        return {"result": _process_memory(pid)}
    if action == "debug_output":     return {"result": _debug_output(int(target) if target else 20)}

    # Databases
    if action == "sqlite_query":
        parts = (target or ",").split(",")
        return {"result": _sqlite_query(parts[0], parts[1] if len(parts) > 1 else "SELECT * FROM sqlite_master")}

    # Remote Monitor
    if action in ("remote_start", "remote_monitor_start"):
        from remote_monitor import get_monitor
        m = get_monitor()
        result = m.start(expose_ngrok=(value != "local"))
        return {"result": result.get("result", str(result))}
    if action in ("remote_stop", "remote_monitor_stop"):
        from remote_monitor import get_monitor
        m = get_monitor()
        return {"result": m.stop().get("result", str(m.stop()))}
    if action == "remote_status":
        from remote_monitor import get_monitor
        m = get_monitor()
        s = m.get_status()
        s.pop("running", None)
        return {"result": json.dumps(s, indent=2)}
    if action == "remote_motion":
        from remote_monitor import get_monitor
        m = get_monitor()
        m.set_motion_detection(True, int(value) if value else 500)
        return {"result": "Motion detection enabled"}
    if action == "remote_motion_off":
        from remote_monitor import get_monitor
        m = get_monitor()
        m.set_motion_detection(False)
        return {"result": "Motion detection disabled"}

    # Keyboard & Mouse Advanced
    if action == "press_hotkey":
        return {"result": _press_hotkey(target or "")}
    if action == "scroll":
        parts = (target or "0,vertical").split(",")
        clicks = int(parts[0]) if parts[0] else 3
        axis = parts[1] if len(parts) > 1 else "vertical"
        return {"result": _scroll(clicks, axis)}
    if action == "type_text":
        return {"result": _type_text(target or "", value)}

    return {"result": f"Unknown action: {action}. Run universal_control with action='sysinfo' to test."}