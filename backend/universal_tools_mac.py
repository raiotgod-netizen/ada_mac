"""
universal_tools_mac.py — Universal PC Control Tool for ADA on macOS

Single tool: universal_control(action, target, value)
Mac equivalents for the Windows universal_tools.

Platform: macOS (Darwin)
"""

from __future__ import annotations
import os
import sys
import subprocess
import base64
import tempfile
import hashlib
import time
import re
import shutil
import platform
from pathlib import Path
from typing import Optional, Dict, Any, List


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _run_shell(cmd: List[str], timeout: int = 10) -> subprocess.CompletedProcess:
    """Run shell command, return result."""
    return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)

def _run_osascript(script: str, timeout: int = 10) -> subprocess.CompletedProcess:
    """Run osascript (AppleScript) command."""
    return subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=timeout)

def _sysctl(key: str) -> str:
    """Get sysctl value."""
    r = _run_shell(["sysctl", "-n", key])
    return r.stdout.strip() if r.returncode == 0 else "N/A"


# ─────────────────────────────────────────────────────────────
# SYSTEM
# ─────────────────────────────────────────────────────────────

def _sysinfo() -> str:
    import psutil, datetime
    try:
        cpu = psutil.cpu_percent(interval=None)
        cpu_count = psutil.cpu_count()
        mem = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        boot = datetime.datetime.fromtimestamp(psutil.boot_time())
        uptime = str(datetime.datetime.now() - boot).split('.')[0]
        model = _sysctl("hw.model")
        macos_version = platform.mac_ver()[0] or "Unknown"
        return (f"Model: {model}\n"
                f"macOS: {macos_version}\n"
                f"CPU: {cpu}% ({cpu_count} cores)\n"
                f"RAM: {mem.percent}% ({mem.used//(1024**3)}/{mem.total//(1024**3)}GB)\n"
                f"Disk: {round(disk.percent,1)}% ({disk.free//(1024**3)}GB free)\n"
                f"Uptime: {uptime}")
    except Exception as e:
        return f"sysinfo error: {e}"

def _cpu_detailed() -> str:
    import psutil, os
    try:
        cpu = psutil.cpu_percent(interval=None)
        per_cpu = psutil.cpu_percent(interval=None, percpu=True)
        load = os.getloadavg() if hasattr(os, 'getloadavg') else [0, 0, 0]
        ctx = psutil.cpu_stats()
        freq = _sysctl("hw.cpufrequency") or _sysctl("hw.cpufrequency_max")
        freq_str = f"{int(freq)/1e9:.2f}GHz" if freq and freq != "N/A" else "N/A"
        return (f"Overall: {cpu}%\nPer core: {per_cpu}\n"
                f"Load (1m): {load[0]} | (5m): {load[1]} | (15m): {load[2]}\n"
                f"Ctx switches: {ctx.ctx_switches:,} | Interrupts: {ctx.interrupts:,}\n"
                f"Frequency: {freq_str}")
    except Exception as e:
        return f"cpu_detailed error: {e}"

def _gpu_info() -> str:
    try:
        result = _run_shell(["system_profiler", "SPDisplaysDataType", "-json"])
        import json
        data = json.loads(result.stdout) if result.returncode == 0 else {}
        gpus = []
        for disp in data.get("SPDisplaysDataType", []):
            gpus.append(f"{disp.get('chipset_model', 'Unknown')} ({disp.get('vram', '?')})")
        return f"GPU: {', '.join(gpus) if gpus else 'Unknown'}"
    except Exception as e:
        return f"gpu_info error: {e}"

def _memory_detailed() -> str:
    import psutil
    try:
        vm = psutil.virtual_memory()
        swap = psutil.swap_memory()
        return (f"Virtual: {vm.percent}% | {vm.used//(1024**3)}GB/{vm.total//(1024**3)}GB\n"
                f"Available: {vm.available//(1024**3)}GB\n"
                f"Swap: {swap.percent}% | {swap.used//(1024**3)}GB/{swap.total//(1024**3)}GB")
    except Exception as e:
        return f"memory_detailed error: {e}"

def _disk_smart() -> str:
    # macOS doesn't expose S.M.A.R.T. via CLI easily — diskutil is the equivalent
    result = _run_shell(["diskutil", "list", "-plist"])
    if result.returncode == 0:
        return f"Use 'diskutil info /dev/disk0' for details. Output: {result.stdout[:400]}"
    return "diskutil not available"

def _thermal_zones() -> str:
    try:
        result = _run_shell(["osx-cpu-temp", "-c"], timeout=3)
        if result.returncode == 0:
            return f"CPU Temp: {result.stdout.strip()}"
    except Exception:
        pass
    # Alternative: powermetrics (requires sudo)
    result = _run_shell(["sh", "-c", "powermetrics --samplers smc | grep -i 'cpu temp' | head -3"], timeout=5)
    if result.returncode == 0 and result.stdout.strip():
        return result.stdout.strip()
    return "Thermal data not available (install 'osx-cpu-temp' or run as root with powermetrics)"


# ─────────────────────────────────────────────────────────────
# NETWORK
# ─────────────────────────────────────────────────────────────

def _net_connections() -> str:
    try:
        import psutil
        lines = []
        for conn in psutil.net_connections(kind='inet')[:50]:
            try:
                lines.append(f"{conn.laddr.ip}:{conn.laddr.port} -> {conn.raddr.ip if conn.raddr else '---'}:{conn.raddr.port if conn.raddr else '---'} [{conn.status}] pid:{conn.pid}")
            except Exception:
                pass
        return "\n".join(lines[:40]) or "No connections"
    except Exception as e:
        return f"net_connections error: {e}"

def _net_arp() -> str:
    result = _run_shell(["arp", "-a"])
    return result.stdout.strip()[:800] if result.returncode == 0 else f"arp error: {result.stderr[:200]}"

def _net_routes() -> str:
    result = _run_shell(["netstat", "-rn"])
    return result.stdout.strip()[:800] if result.returncode == 0 else f"routes error: {result.stderr[:200]}"

def _net_dns_cache() -> str:
    result = _run_shell(["scutil", "--dns"])
    return result.stdout.strip()[:600] if result.returncode == 0 else f"dns cache error: {result.stderr[:200]}"

def _wifi_networks() -> str:
    result = _run_shell(["/System/Library/PrivateFrameworks/Apple80211.framework/Versions/A/Resources/airport", "-s"])
    return result.stdout.strip()[:600] if result.returncode == 0 else f"wifi error: {result.stderr[:200]}"

def _wifi_connected() -> str:
    result = _run_shell(["networksetup", "-getairportnetwork", "en0"])
    return result.stdout.strip()[:400] if result.returncode == 0 else f"wifi error: {result.stderr[:200]}"

def _firewall_rules() -> str:
    result = _run_shell(["/usr/libexec/ApplicationFirewall/socketfilterfw", "--listapps"])
    return result.stdout.strip()[:800] if result.returncode == 0 else f"firewall error: {result.stderr[:200]}"


# ─────────────────────────────────────────────────────────────
# WINDOWS (macOS = App Windows via osascript)
# ─────────────────────────────────────────────────────────────

def _windows_full() -> str:
    """List open windows via osascript."""
    script = """
    tell application "System Events"
        set windowList to {}
        set appList to every application process whose frontmost is true
        repeat with theApp in appList
            set appName to name of theApp
            try
                set winList to every window of theApp
                repeat with theWin in winList
                    set winName to name of theWin
                    if winName is not "" then
                        set end of windowList to "[" & appName & "] " & winName
                    end if
                end repeat
            end try
        end repeat
    end tell
    return windowList as string
    """
    r = _run_osascript(script)
    if r.returncode == 0 and r.stdout.strip():
        return r.stdout.strip()
    # Fallback: list frontmost apps
    r2 = _run_osascript('tell application "System Events" to get name of every application process whose visible is true')
    return f"Open apps: {r2.stdout.strip()}" if r2.returncode == 0 else "No window data"

def _window_tree() -> str:
    return _windows_full()  # macOS doesn't have window hierarchy the same way

def _window_info(target: str) -> str:
    script = f'''
    tell application "System Events"
        set appList to every application process whose name contains "{target}"
        repeat with theApp in appList
            set appName to name of theApp
            set winList to every window of theApp
            if (count of winList) > 0 then
                repeat with theWin in winList
                    set winName to name of theWin
                    return "App: " & appName & " | Window: " & winName
                end repeat
            end if
        end repeat
    end tell
    '''
    r = _run_osascript(script)
    return r.stdout.strip() if r.returncode == 0 else f"No window matching: {target}"

def _window_action(target: str, action: str) -> str:
    """Close/minimize/maximize/focus a window by name."""
    app_script = f'''
    tell application "System Events"
        set targetApp to first application process whose name contains "{target}"
        set frontmost of targetApp to true
    end tell
    '''
    if action == "close":
        script = f'''
        tell application "System Events"
            set targetApp to first application process whose name contains "{target}"
            set winList to every window of targetApp
            if (count of winList) > 0 then
                perform close of every window of targetApp
            end if
        end tell
        '''
    elif action == "minimize":
        script = f'''
        tell application "System Events"
            set targetApp to first application process whose name contains "{target}"
            set winList to every window of targetApp
            if (count of winList) > 0 then
                perform miniaturize of every window of targetApp
            end if
        end tell
        '''
    elif action == "maximize":
        return f"maximize not directly supported on macOS — use fullscreen (Ctrl+Cmd+F)"
    elif action == "focus":
        script = app_script
    else:
        return f"Unknown action: {action}"

    r = _run_osascript(script)
    return f"{action} on {target}: {'OK' if r.returncode == 0 else r.stderr[:200]}"

def _window_set_pos(target: str, x: int, y: int, w: int = None, h: int = None) -> str:
    # macOS doesn't allow arbitrary window positioning via AppleScript
    return f"window_set_pos not supported on macOS (x={x}, y={y})"

def _window_always_on_top(target: str, enable: bool = True) -> str:
    return "Always-on-top not natively supported on macOS"


# ─────────────────────────────────────────────────────────────
# PROCESSES
# ─────────────────────────────────────────────────────────────

def _processes_detailed() -> str:
    try:
        import psutil
        lines = []
        for p in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_percent', 'num_threads']):
            try:
                info = p.info
                lines.append(f"PID:{info['pid']} {info['name'][:30]:30} CPU:{info['cpu_percent']}% MEM:{round(info['memory_percent'],1)}% TH:{info['num_threads']}")
            except Exception:
                pass
        return "\n".join(lines[:40])
    except Exception as e:
        return f"processes error: {e}"

def _process_tree() -> str:
    try:
        import psutil
        lines = []
        for p in psutil.process_iter(['pid', 'name', 'ppid']):
            try:
                info = p.info
                lines.append(f"PID:{info['pid']} PPID:{info['ppid']} {info['name']}")
            except Exception:
                pass
        return "\n".join(lines[:60])
    except Exception as e:
        return f"process_tree error: {e}"

def _process_kill(target: str) -> str:
    try:
        import psutil
        killed = []
        for p in psutil.process_iter(['pid', 'name']):
            try:
                if target.lower() in p.info['name'].lower():
                    pid = p.info['pid']
                    subprocess.run(['kill', str(pid)], capture_output=True)
                    killed.append(f"{p.info['name']} (PID:{pid})")
            except Exception:
                pass
        return f"Killed: {', '.join(killed) if killed else 'none found'}"
    except Exception as e:
        return f"process_kill error: {e}"

def _process_kill_tree(target: str) -> str:
    # macOS kill -9 equivalent
    return _process_kill(target) + " (tree mode)"


# ─────────────────────────────────────────────────────────────
# FILES
# ─────────────────────────────────────────────────────────────

def _dir_tree(path: str = None, depth: int = 3) -> str:
    path = path or os.path.expanduser("~")
    try:
        lines = []
        for root, dirs, files in os.walk(path):
            depth_dir = root.replace(path, '').count(os.sep)
            if depth_dir >= depth:
                dirs.clear()
                continue
            lines.append(f"{'  '*depth_dir}{os.path.basename(root)}/")
            for f in files[:5]:
                lines.append(f"{'  '*(depth_dir+1)}{f}")
        return "\n".join(lines[:80])
    except Exception as e:
        return f"dir_tree error: {e}"

def _file_search(pattern: str, path: str = None) -> str:
    path = path or os.path.expanduser("~")
    try:
        result = _run_shell(["find", path, "-iname", f"*{pattern}*", "-maxdepth", "5", "-type", "f"])
        matches = result.stdout.strip().split("\n")[:30]
        return "\n".join(m for m in matches if m)
    except Exception as e:
        return f"file_search error: {e}"

def _file_info(path: str) -> str:
    try:
        p = Path(path)
        if not p.exists():
            return f"File not found: {path}"
        stat = p.stat()
        return (f"Path: {p.absolute()}\n"
                f"Size: {stat.st_size:,} bytes\n"
                f"Created: {time.ctime(stat.st_ctime)}\n"
                f"Modified: {time.ctime(stat.st_mtime)}\n"
                f"Accessed: {time.ctime(stat.st_atime)}")
    except Exception as e:
        return f"file_info error: {e}"

def _file_duplicate_finder(path: str = None) -> str:
    path = path or os.path.expanduser("~")
    try:
        hashes = {}
        dups = []
        for root, dirs, files in os.walk(path):
            for f in files:
                fp = os.path.join(root, f)
                try:
                    size = os.path.getsize(fp)
                    if size < 1024:
                        continue
                    with open(fp, 'rb') as fh:
                        h = hashlib.md5(fh.read(4096)).hexdigest()
                    key = (size, h)
                    if key in hashes:
                        dups.append(f"{fp} == {hashes[key]}")
                    else:
                        hashes[key] = fp
                except Exception:
                    pass
                if len(dups) >= 20:
                    break
            if len(dups) >= 20:
                break
        return "\n".join(dups) if dups else "No duplicates found"
    except Exception as e:
        return f"file_duplicate_finder error: {e}"

def _file_watch(path: str, seconds: int = 5) -> str:
    try:
        import watchdog
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
        class Handler(FileSystemEventHandler):
            def __init__(self):
                self.events = []
            def on_any_event(self, event):
                if not event.is_directory:
                    self.events.append(f"{event.event_type}: {event.src_path}")
        h = Handler()
        Observer().schedule(h, path, recursive=True).start()
        time.sleep(seconds)
        Observer().stop()
        return "\n".join(h.events[:20]) if h.events else "No changes detected"
    except Exception as e:
        return f"file_watch error: {e}"

def _file_delete(path: str) -> str:
    try:
        p = Path(path)
        if p.is_file():
            p.unlink()
            return f"Deleted: {path}"
        elif p.is_dir():
            shutil.rmtree(p)
            return f"Deleted dir: {path}"
        return f"Not found: {path}"
    except Exception as e:
        return f"file_delete error: {e}"


# ─────────────────────────────────────────────────────────────
# SCREENSHOT & VISION
# ─────────────────────────────────────────────────────────────

def _screenshot() -> str:
    try:
        import mss
        with mss.mss() as sct:
            monitor = sct.monitors[1]
            screenshot = sct.grab(monitor)
            img_bytes = mss.tools.to_png(screenshot.rgb, screenshot.size)
            b64 = base64.b64encode(img_bytes).decode('utf-8')
            return f"[SCREENSHOT_BASE64:{len(b64)} chars]"
    except Exception:
        # Fallback: screencapture command
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        tmp.close()
        r = _run_shell(["screencapture", "-x", tmp.name])
        if r.returncode == 0 and Path(tmp.name).exists():
            b64 = base64.b64encode(Path(tmp.name).read_bytes()).decode('utf-8')
            os.unlink(tmp.name)
            return f"[SCREENSHOT_BASE64:{len(b64)} chars]"
        return f"screenshot error: {r.stderr[:200]}"

def _screenshot_area(x: int, y: int, w: int, h: int) -> str:
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp.close()
    r = _run_shell(["screencapture", "-x", "-R", f"{x},{y},{w},{h}", tmp.name])
    if r.returncode == 0 and Path(tmp.name).exists():
        b64 = base64.b64encode(Path(tmp.name).read_bytes()).decode('utf-8')
        os.unlink(tmp.name)
        return f"[SCREENSHOT_AREA_BASE64:{len(b64)} chars]"
    return f"screenshot_area error: {r.stderr[:200]}"

def _screen_regions() -> str:
    r = _run_shell(["system_profiler", "SPDisplaysDataType", "-json"])
    return r.stdout[:400] if r.returncode == 0 else "No display info"

def _ocr_screen() -> str:
    try:
        import mss
        with mss.mss() as sct:
            monitor = sct.monitors[1]
            screenshot = sct.grab(monitor)
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
                f.write(mss.tools.to_png(screenshot.rgb, screenshot.size))
                temp_path = f.name
        from ocr_engine import OCREngine
        ocr = OCREngine()
        result = ocr.extract(temp_path)
        os.unlink(temp_path)
        if result.get("ocr_ready"):
            lines = [l['text'] for l in result.get("lines", []) if l.get('text', '').strip()]
            return " | ".join(lines[:30])
        return "No text found"
    except Exception as e:
        return f"ocr_screen error: {e}"


# ─────────────────────────────────────────────────────────────
# CLIPBOARD & INPUT
# ─────────────────────────────────────────────────────────────

def _clipboard_read() -> str:
    try:
        r = _run_shell(["pbpaste"])
        return r.stdout[:500] if r.returncode == 0 else "(clipboard empty)"
    except Exception as e:
        return f"clipboard_read error: {e}"

def _clipboard_write(text: str) -> str:
    try:
        proc = subprocess.Popen(["pbcopy"], stdin=subprocess.PIPE)
        proc.communicate(input=text.encode('utf-8'))
        return f"Written to clipboard ({len(text)} chars)"
    except Exception as e:
        return f"clipboard_write error: {e}"

def _key_send(keys: str) -> str:
    try:
        import pyautogui
        pyautogui.write(keys, interval=0.02)
        return f"Sent keys: {keys[:50]}"
    except Exception as e:
        return f"key_send error: {e}"

def _mouse_move(x: int, y: int) -> str:
    try:
        import pyautogui
        pyautogui.moveTo(x, y, duration=0.2)
        return f"Moved mouse to ({x},{y})"
    except Exception as e:
        return f"mouse_move error: {e}"

def _mouse_click(x: int = None, y: int = None, button: str = "left") -> str:
    try:
        import pyautogui
        if x is not None and y is not None:
            pyautogui.click(x, y, button=button)
        else:
            pyautogui.click(button=button)
        return f"Clicked {button} at ({x},{y})" if x else f"Clicked {button}"
    except Exception as e:
        return f"mouse_click error: {e}"

def _mouse_drag(x1: int, y1: int, x2: int, y2: int) -> str:
    try:
        import pyautogui
        pyautogui.moveTo(x1, y1)
        pyautogui.drag(x2-x1, y2-y1, duration=0.5)
        return f"Dragged from ({x1},{y1}) to ({x2},{y2})"
    except Exception as e:
        return f"mouse_drag error: {e}"

def _press_hotkey(keys: str) -> str:
    try:
        import pyautogui
        key_list = [k.strip().lower() for k in keys.split("+")]
        pyautogui.hotkey(*key_list)
        return f"Pressed hotkey: {'+'.join(key_list)}"
    except Exception as e:
        return f"press_hotkey error: {e}"

def _scroll(clicks: int, axis: str = "vertical") -> str:
    try:
        import pyautogui
        if axis == "horizontal" or axis == "h":
            pyautogui.hscroll(clicks)
        else:
            pyautogui.scroll(clicks)
        return f"Scrolled {axis} {clicks} clicks"
    except Exception as e:
        return f"scroll error: {e}"

def _type_text(text: str, interval: float = None) -> str:
    try:
        import pyautogui
        pyautogui.write(text, interval=interval or 0.02)
        return f"Typed {len(text)} characters"
    except Exception as e:
        return f"type_text error: {e}"


# ─────────────────────────────────────────────────────────────
# APP CONTROL
# ─────────────────────────────────────────────────────────────

def _app_open(target: str) -> str:
    try:
        if target.startswith('http://') or target.startswith('https://'):
            subprocess.Popen(["open", target])
            return f"Opened URL: {target}"
        if os.path.exists(target):
            subprocess.Popen(["open", target])
            return f"Opened: {target}"
        # Try by app name
        subprocess.Popen(["open", "-a", target])
        return f"Opened app: {target}"
    except Exception as e:
        return f"app_open error: {e}"

def _app_close(target: str) -> str:
    try:
        r = _run_osascript(f'tell application "{target}" to quit')
        if r.returncode == 0:
            return f"Closed: {target}"
        # Fallback: kill by name
        r2 = _run_shell(["pkill", "-f", target])
        return f"Closed (pkill): {target}" if r2.returncode == 0 else f"Close failed: {r2.stderr[:200]}"
    except Exception as e:
        return f"app_close error: {e}"


# ─────────────────────────────────────────────────────────────
# SERVICES
# ─────────────────────────────────────────────────────────────

def _service_control(service_name: str, action: str) -> str:
    # macOS uses launchd — not directly equivalent
    return f"launchd services: use 'launchctl {action}' for {service_name}"

def _startup_list() -> str:
    result = _run_shell(["launchctl", "list"])
    return result.stdout.strip()[:800] if result.returncode == 0 else f"startup_list error: {result.stderr[:200]}"


# ─────────────────────────────────────────────────────────────
# ENVIRONMENT
# ─────────────────────────────────────────────────────────────

def _env_get(var: str = None) -> str:
    if var:
        return os.environ.get(var, "(not set)")
    return "\n".join([f"{k}={v}" for k, v in os.environ.items()][:30])

def _env_set(var: str, value: str) -> str:
    try:
        os.environ[var] = value
        return f"Set {var}={value} (current process only)"
    except Exception as e:
        return f"env_set error: {e}"


# ─────────────────────────────────────────────────────────────
# PC CONTROL (macOS)
# ─────────────────────────────────────────────────────────────

def _reboot() -> str:
    subprocess.run(["shutdown", "-r", "+1"], capture_output=True)
    return "Reboot in 1 minute"

def _shutdown() -> str:
    subprocess.run(["shutdown", "-h", "+1"], capture_output=True)
    return "Shutdown in 1 minute"

def _hibernate() -> str:
    return "Hibernate not directly available on macOS — use 'pmset sleep'"

def _lock() -> str:
    # macOS lock = start screen saver or lock via peck
    subprocess.Popen(["/System/Library/CoreServices/Menu Extras/User.menu/Contents/Resources/CGSession", "-suspend"])
    return "Workstation lock initiated"


# ─────────────────────────────────────────────────────────────
# SYSTEM LOGS & STATUS
# ─────────────────────────────────────────────────────────────

def _system_logs(lines: int = 20) -> str:
    result = _run_shell(["log", "show", "--predicate", "process == 'loginwindow'", "--last", f"{lines}m"])
    return result.stdout.strip()[:800] if result.returncode == 0 else f"system_logs error: {result.stderr[:200]}"

def _crash_dump_check() -> str:
    crash_path = "/Library/Logs/DiagnosticReports"
    if os.path.exists(crash_path):
        files = sorted(Path(crash_path).glob("*.crash"))[-5:]
        total = sum(f.stat().st_size for f in files)
        return f"Crash reports: {len(files)} recent, {total//(1024*1024)}MB in {crash_path}"
    return "No crash reports found"

def _update_status() -> str:
    result = _run_shell(["softwareupdate", "-l"])
    return result.stdout.strip()[:400] if result.returncode == 0 else "No updates available"


# ─────────────────────────────────────────────────────────────
# EXCEL (portable — no change needed)
# ─────────────────────────────────────────────────────────────

def _excel_read(path: str, sheet: str = None, cell_range: str = None) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        if cell_range:
            data = [[str(cell.value or '') for cell in row] for row in ws[cell_range]]
        else:
            data = [[str(cell.value or '') for cell in row] for row in ws.iter_rows(max_row=20, max_col=10)]
        return "\n".join([" | ".join(row) for row in data])
    except Exception as e:
        return f"excel_read error: {e}"

def _excel_write(path: str, cell: str, value: str, sheet: str = None) -> str:
    try:
        import openpyxl
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
        ws = wb[sheet] if sheet else wb.active
        ws[cell] = value
        wb.save(path)
        return f"Written {value} to {cell} in {path}"
    except Exception as e:
        return f"excel_write error: {e}"


# ─────────────────────────────────────────────────────────────
# BROWSER TABS (macOS — via osascript)
# ─────────────────────────────────────────────────────────────

def _browser_tabs() -> str:
    browsers = ["Safari", "Google Chrome", "Firefox"]
    result_lines = []
    for browser in browsers:
        script = f'''
        tell application "{browser}"
            set tabList to {{}}
            try
                set winList to every window
                repeat with theWin in winList
                    set tabCount to count of tabs of theWin
                    set end of tabList to "{browser}: " & tabCount & " tabs"
                end repeat
            end try
        end tell
        '''
        r = _run_osascript(script)
        if r.returncode == 0 and r.stdout.strip():
            result_lines.append(r.stdout.strip())
    return "\n".join(result_lines) if result_lines else "No browser tabs found"


# ─────────────────────────────────────────────────────────────
# HARDWARE & SENSORS
# ─────────────────────────────────────────────────────────────

def _fan_speed() -> str:
    try:
        result = _run_shell(["powermetrics", "--samplers", "smc", "--iterations", "1"], timeout=10)
        for line in result.stdout.split("\n"):
            if "fan" in line.lower():
                return line.strip()
        return "Fan speed not available (requires powermetrics as root)"
    except Exception as e:
        return f"fan_speed error: {e}"

def _disk_partitions() -> str:
    result = _run_shell(["diskutil", "list"])
    return result.stdout.strip()[:800] if result.returncode == 0 else f"disk_partitions error: {result.stderr[:200]}"

def _hardware_usb() -> str:
    result = _run_shell(["system_profiler", "SPUSBDataType", "-json"])
    return result.stdout.strip()[:800] if result.returncode == 0 else f"hardware_usb error: {result.stderr[:200]}"

def _bluetooth_devices() -> str:
    result = _run_shell(["system_profiler", "SPBluetoothDataType", "-json"])
    return result.stdout.strip()[:600] if result.returncode == 0 else f"bluetooth_devices error: {result.stderr[:200]}"


# ─────────────────────────────────────────────────────────────
# NETWORK ADVANCED
# ─────────────────────────────────────────────────────────────

def _ping_status(host: str) -> str:
    result = _run_shell(["ping", "-c", "3", host])
    if result.returncode == 0:
        lines = result.stdout.split("\n")
        for line in lines:
            if "round-trip" in line or "avg" in line:
                return line.strip()
        return result.stdout.strip()[:200]
    return f"ping error: {result.stderr[:200]}"

def _port_scan(port_range: str = "1-1024") -> str:
    return "Use 'nmap' (brew install nmap) for port scanning on macOS"

def _network_map() -> str:
    result = _run_shell(["arp", "-a"])
    return result.stdout.strip()[:800] if result.returncode == 0 else f"network_map error: {result.stderr[:200]}"


# ─────────────────────────────────────────────────────────────
# APPLICATIONS ADVANCED
# ─────────────────────────────────────────────────────────────

def _pdf_read(path: str) -> str:
    try:
        import PyPDF2
        with open(path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            text = ""
            for page in reader.pages[:5]:
                text += page.extract_text() or ""
            return text[:500] if text.strip() else "No text found in PDF"
    except ImportError:
        return "PyPDF2 not installed"
    except Exception as e:
        return f"pdf_read error: {e}"

def _compress_file(source: str, destination: str = None) -> str:
    try:
        if not destination:
            destination = source + ".zip"
        subprocess.run(["zip", "-r", destination, source], capture_output=True)
        return f"Compressed: {destination}"
    except Exception as e:
        return f"compress_file error: {e}"

def _uncompress_file(source: str, destination: str = None) -> str:
    try:
        if destination is None:
            destination = source.replace('.zip', '_unzipped')
        subprocess.run(["unzip", "-o", source, "-d", destination], capture_output=True)
        return f"Extracted to: {destination}"
    except Exception as e:
        return f"uncompress_file error: {e}"


# ─────────────────────────────────────────────────────────────
# SECURITY
# ─────────────────────────────────────────────────────────────

def _firewall_status() -> str:
    r = _run_shell(["/usr/libexec/ApplicationFirewall/socketfilterfw", "--getglobalstate"])
    return r.stdout.strip() if r.returncode == 0 else "Firewall status unknown"

def _antivirus_status() -> str:
    # macOS doesn't have built-in antivirus status like Windows Defender
    return "Use 'which knockknock' or 'which littoral' for malware checks"

def _uac_check() -> str:
    return "UAC is a Windows feature — macOS uses Gatekeeper and System Integrity Protection"


# ─────────────────────────────────────────────────────────────
# ADVANCED DEBUG
# ─────────────────────────────────────────────────────────────

def _process_memory(pid: int = None) -> str:
    try:
        import psutil
        if pid:
            p = psutil.Process(pid)
            return f"PID {pid}: {p.name()} - MEM: {p.memory_info().rss//(1024**2)}MB"
        lines = []
        for p in psutil.process_iter(['pid', 'name', 'memory_info']):
            try:
                info = p.info
                lines.append(f"PID:{info['pid']} {info['name'][:30]:30} MEM:{info['memory_info'].rss//(1024**2):6}MB")
            except Exception:
                pass
        return "\n".join(lines[:30])
    except Exception as e:
        return f"process_memory error: {e}"

def _debug_output(lines: int = 20) -> str:
    result = _run_shell(["log", "show", "--predicate", "eventMessage contains 'error'", "--last", f"{lines}m"])
    return result.stdout.strip()[:1000] if result.returncode == 0 else f"debug_output error: {result.stderr[:200]}"


# ─────────────────────────────────────────────────────────────
# MAIN DISPATCHER
# ─────────────────────────────────────────────────────────────

def universal_control(action: str, target: str = None, value: str = None) -> Dict[str, Any]:
    """
    Universal PC control tool — macOS edition.
    Action: The action to perform
    Target: Target entity (window name, file path, etc.)
    Value: Additional parameter
    """
    action = action.lower().strip() if action else ""

    # System
    if action == "sysinfo":          return {"result": _sysinfo()}
    if action == "cpu_detailed":    return {"result": _cpu_detailed()}
    if action == "gpu_info":        return {"result": _gpu_info()}
    if action == "memory_detailed": return {"result": _memory_detailed()}
    if action == "disk_smart":     return {"result": _disk_smart()}
    if action == "thermal_zones":   return {"result": _thermal_zones()}

    # Network
    if action == "net_connections": return {"result": _net_connections()}
    if action == "net_arp":         return {"result": _net_arp()}
    if action == "net_routes":      return {"result": _net_routes()}
    if action == "net_dns_cache":   return {"result": _net_dns_cache()}
    if action == "wifi_networks":   return {"result": _wifi_networks()}
    if action == "wifi_connected":  return {"result": _wifi_connected()}
    if action == "firewall_rules":  return {"result": _firewall_rules()}

    # Windows / App Windows
    if action == "windows_full":          return {"result": _windows_full()}
    if action == "window_tree":          return {"result": _window_tree()}
    if action == "window_info":          return {"result": _window_info(target or "")}
    if action in ("window_close", "window_minimize", "window_maximize", "window_focus"):
        return {"result": _window_action(target or "", action.split("_")[1])}
    if action == "window_set_pos":
        return {"result": _window_set_pos(target or "", 0, 0)}
    if action == "window_always_on_top":
        return {"result": _window_always_on_top(target or "", value == "true")}

    # Processes
    if action == "processes_detailed": return {"result": _processes_detailed()}
    if action == "process_tree":       return {"result": _process_tree()}
    if action == "process_kill":      return {"result": _process_kill(target or "")}
    if action == "process_kill_tree": return {"result": _process_kill_tree(target or "")}

    # Files
    if action == "dir_tree":          return {"result": _dir_tree(target)}
    if action == "file_search":       return {"result": _file_search(target or "~")}
    if action == "file_info":         return {"result": _file_info(target or "")}
    if action == "file_duplicate_finder": return {"result": _file_duplicate_finder(target)}
    if action == "file_watch":         return {"result": _file_watch(target or os.path.expanduser("~"), int(value) if value else 5)}
    if action == "file_delete":       return {"result": _file_delete(target or "")}

    # Screenshot & Vision
    if action == "screenshot":       return {"result": _screenshot()}
    if action == "screenshot_area":
        parts = (target or "0,0,100,100").split(",")
        x, y, w, h = int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3])
        return {"result": _screenshot_area(x, y, w, h)}
    if action == "screen_regions":   return {"result": _screen_regions()}
    if action == "ocr_screen":       return {"result": _ocr_screen()}

    # Clipboard & Input
    if action == "clipboard_read":   return {"result": _clipboard_read()}
    if action == "clipboard_write":  return {"result": _clipboard_write(target or "")}
    if action == "key_send":         return {"result": _key_send(target or "")}
    if action == "mouse_move":
        parts = (target or "0,0").split(",")
        return {"result": _mouse_move(int(parts[0]), int(parts[1]))}
    if action == "mouse_click":
        parts = (target or ",").split(",")
        x = int(parts[0]) if parts[0] else None
        y = int(parts[1]) if len(parts) > 1 and parts[1] else None
        return {"result": _mouse_click(x, y, value or "left")}
    if action == "mouse_drag":
        parts = (target or "0,0,0,0").split(",")
        return {"result": _mouse_drag(int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3]))}

    # App Control
    if action == "app_open":         return {"result": _app_open(target or "")}
    if action == "app_close":        return {"result": _app_close(target or "")}

    # Services
    if action == "service_control":
        parts = (target or ",").split(",")
        return {"result": _service_control(parts[0], parts[1] if len(parts) > 1 else "start")}
    if action == "startup_list":     return {"result": _startup_list()}

    # Environment
    if action == "env_get":          return {"result": _env_get(target)}
    if action == "env_set":          return {"result": _env_set(target or "", value or "")}
    if action == "registry_read":    return {"result": "Registry not available on macOS"}
    if action == "registry_write":   return {"result": "Registry not available on macOS"}

    # PC Control
    if action == "reboot":           return {"result": _reboot()}
    if action == "shutdown":         return {"result": _shutdown()}
    if action == "hibernate":        return {"result": _hibernate()}
    if action == "lock":             return {"result": _lock()}

    # System Logs
    if action == "system_logs":      return {"result": _system_logs(int(target) if target else 20)}
    if action == "crash_dump_check": return {"result": _crash_dump_check()}
    if action == "update_status":    return {"result": _update_status()}

    # Excel
    if action == "excel_read":       return {"result": _excel_read(target or "", value)}
    if action == "excel_write":
        parts = (target or "").split(",")
        return {"result": _excel_write(parts[0], parts[1] if len(parts) > 1 else "A1", value or "")}

    # Browser
    if action == "browser_tabs":     return {"result": _browser_tabs()}

    # Hardware & Sensors
    if action == "fan_speed":        return {"result": _fan_speed()}
    if action == "disk_partitions": return {"result": _disk_partitions()}
    if action == "hardware_usb":     return {"result": _hardware_usb()}
    if action == "bluetooth_devices":return {"result": _bluetooth_devices()}

    # Network Advanced
    if action == "ping_status":      return {"result": _ping_status(target or "google.com")}
    if action == "port_scan":       return {"result": _port_scan(target or "1-1024")}
    if action == "network_map":      return {"result": _network_map()}

    # Applications Advanced
    if action == "pdf_read":         return {"result": _pdf_read(target or "")}
    if action == "compress_file":     return {"result": _compress_file(target or "", value)}
    if action == "uncompress_file":  return {"result": _uncompress_file(target or "", value)}

    # Security
    if action == "firewall_status":   return {"result": _firewall_status()}
    if action == "antivirus_status": return {"result": _antivirus_status()}
    if action == "uac_check":         return {"result": _uac_check()}

    # Advanced Debug
    if action == "process_memory":
        pid = int(target) if target and target.isdigit() else None
        return {"result": _process_memory(pid)}
    if action == "debug_output":     return {"result": _debug_output(int(target) if target else 20)}

    # Keyboard & Mouse Advanced
    if action == "press_hotkey":
        return {"result": _press_hotkey(target or "")}
    if action == "scroll":
        parts = (target or "0,vertical").split(",")
        clicks = int(parts[0]) if parts[0] else 3
        axis = parts[1] if len(parts) > 1 else "vertical"
        return {"result": _scroll(clicks, axis)}
    if action == "type_text":
        return {"result": _type_text(target or "")}

    return {"result": f"Unknown action: {action}. Run universal_control with action='sysinfo' to test."}
