"""
Creates Formulario_Inscripcion.xlsx with all sheets and data.
The VBA form and macros are in separate files to import manually.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

ADA_DIR = r'C:\Users\raiot\OneDrive\Escritorio\ADA\ada_v2'
OUTPUT_XLSX = os.path.join(ADA_DIR, 'Formulario_Inscripcion.xlsx')

wb = openpyxl.Workbook()

# === Hoja Apoyo ===
ws = wb.active
ws.title = 'Apoyo'
headers = ['IDCurso', 'NombreCurso', 'Precio']
for col, h in enumerate(headers, 1):
    c = ws.cell(1, col)
    c.value = h
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='4472C4')
    c.alignment = Alignment(horizontal='center')

cursos = [
    ('CUR001', 'Introduccion a la Programacion', 15000),
    ('CUR002', 'Python Avanzado', 25000),
    ('CUR003', 'Excel con Macros', 18000),
    ('CUR004', 'Analisis de Datos con Excel', 22000),
    ('CUR005', 'Power BI Fundamentals', 30000),
    ('CUR006', 'Machine Learning Basics', 45000),
    ('CUR007', 'Base de Datos SQL', 20000),
    ('CUR008', 'Presentaciones Ejecutivas', 12000),
]
for i, (cod, nom, pre) in enumerate(cursos, 2):
    ws.cell(i, 1).value = cod
    ws.cell(i, 2).value = nom
    ws.cell(i, 3).value = pre
    ws.cell(i, 3).number_format = '"$"#,##0.00'

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15

# === Hoja Datos ===
ws_datos = wb.create_sheet('Datos')
hd = ['Nombre y Apellido', 'ID Curso', 'Precio', 'Fecha Inscripcion']
for col, h in enumerate(hd, 1):
    c = ws_datos.cell(1, col)
    c.value = h
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='2E75B6')
    c.alignment = Alignment(horizontal='center')

ws_datos.column_dimensions['A'].width = 30
ws_datos.column_dimensions['B'].width = 12
ws_datos.column_dimensions['C'].width = 15
ws_datos.column_dimensions['D'].width = 22

# === Hoja Menu ===
ws_menu = wb.create_sheet('Menu')
ws_menu['A1'] = 'FORMULARIO DE INSCRIPCION'
ws_menu['A1'].font = Font(bold=True, size=16)
ws_menu['A3'] = 'Para usar el formulario:'
ws_menu['A4'] = '1. Presione Alt+F11 para abrir el editor VBA'
ws_menu['A5'] = '2. Importe el archivo UserForm1.frm (Insertar > Importar archivo)'
ws_menu['A6'] = '3. Importe el archivo ModuloVBA.bas (Insertar > Importar archivo)'
ws_menu['A7'] = '4. Cierre el editor VBA y presione el boton Macro abajo'
ws_menu['A9'] = '5. Asigne la macro "MostrarFormulario" al boton'
ws_menu['A11'] = 'Haga clic en: Desarrollador > Insertar > Boton (control de formulario)'
ws_menu['A12'] = 'Luego asigne la macro "MostrarFormulario" al boton creado'
ws_menu['A1'].font = Font(bold=True, size=16, color='1F4E79')

# Add button on menu sheet

wb.save(OUTPUT_XLSX)
print(f'Guardado: {OUTPUT_XLSX}')
