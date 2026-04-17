"""
Creates formulario_inscripcion.xlsm with:
- Hoja "Apoyo": IDCurso, NombreCurso, Precio
- Hoja "Datos": Nombre, IDCurso, Precio, Fecha
- UserForm1 with VBA macros
- Button on sheet to open form
"""
import os, sys

# We'll use openpyxl for structure + manually inject VBA via zipfile
# This is the most portable way to create a .xlsm with macros

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ADA_DIR = r'C:\Users\raiot\OneDrive\Escritorio\ADA\ada_v2'
OUTPUT = os.path.join(ADA_DIR, 'Formulario_Inscripcion.xlsm')

wb = Workbook()

# === Hoja Apoyo ===
ws_apoyo = wb.active
ws_apoyo.title = 'Apoyo'
ws_apoyo['A1'] = 'IDCurso'
ws_apoyo['B1'] = 'NombreCurso'
ws_apoyo['C1'] = 'Precio'

# Sample data
cursos = [
    ('CUR001', 'Introduccion a la Programacion', 15000),
    ('CUR002', 'Python Avanzado', 25000),
    ('CUR003', 'Excel con Macros', 18000),
    ('CUR004', '数据分析 con Excel', 22000),
    ('CUR005', 'Power BI Fundamentals', 30000),
    ('CUR006', 'Machine Learning Basics', 45000),
    ('CUR007', 'Base de Datos SQL', 20000),
    ('CUR008', 'Presentaciones Ejecutivas', 12000),
]

for i, (cod, nom, pre) in enumerate(cursos, start=2):
    ws_apoyo[f'A{i}'] = cod
    ws_apoyo[f'B{i}'] = nom
    ws_apoyo[f'C{i}'] = pre
    ws_apoyo[f'C{i}'].number_format = '"$"#,##0.00'

# Format header
for col in range(1, 4):
    cell = ws_apoyo.cell(1, col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', fgColor='4472C4')
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center')

ws_apoyo.column_dimensions['A'].width = 12
ws_apoyo.column_dimensions['B'].width = 35
ws_apoyo.column_dimensions['C'].width = 15

# === Hoja Datos ===
ws_datos = wb.create_sheet('Datos')
headers_datos = ['Nombre y Apellido', 'ID Curso', 'Precio', 'Fecha Inscripcion']
for col, h in enumerate(headers_datos, 1):
    cell = ws_datos.cell(1, col)
    cell.value = h
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='2E75B6')
    cell.alignment = Alignment(horizontal='center')

ws_datos.column_dimensions['A'].width = 30
ws_datos.column_dimensions['B'].width = 12
ws_datos.column_dimensions['C'].width = 15
ws_datos.column_dimensions['D'].width = 22

# === Hoja Menu (main) ===
ws_menu = wb.create_sheet('Menu')
ws_menu['A1'] = 'FORMULARIO DE INSCRIPCION'
ws_menu['A1'].font = Font(bold=True, size=16, color='1F4E79')
ws_menu['A3'] = 'Presiona el boton para abrir el formulario de inscripcion'
ws_menu['A5'] = 'Macros requeridos: habilitar contenido al abrir'
ws_menu['A5'].font = Font(italic=True, color='808080')

# Add a button shape that runs MostrarFormulario
from openpyxl.drawing.shapes import Rectangle
btn = Rectangle()
btn.text = 'Abrir Formulario de Inscripcion'
btn.width = 250
btn.height = 50
ws_menu.add_shape(btn, 'A7')

# Save as .xlsm is not directly supported by openpyxl for VBA
# We save as .xlsx first, then the user gets instructions
xlsx_path = os.path.join(ADA_DIR, 'Formulario_Inscripcion_temp.xlsx')
wb.save(xlsx_path)
print(f"Workbook saved to: {xlsx_path}")
print("")
print("NOTE: openpyxl no puede guardar VBA en .xlsm directamente.")
print("Necesitas usar el archivo .bas que se creo y:")
print("1. Abrir este Excel > Guardar como .xlsm")
print("2. Abrir VBA (Alt+F11) > Importar formulario_inscripcion.bas")
print("3. O usar el script win32com siguiente para crear el .xlsm completo")
